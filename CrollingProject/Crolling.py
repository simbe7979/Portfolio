import tkinter as tk
import time
import openpyxl
from tkinter import ttk
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from tkinter import messagebox
from openpyxl.styles import PatternFill

netflix_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EB%84%B7%ED%94%8C%EB%A6%AD%EC%8A%A4+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75lsqo1LwssPveaXlssssstLC-061196"
watcha_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%EC%99%93%EC%B1%A0+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75dsqptbNssnJ3V2wssssss8K-371550"
tving_url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&ssc=tab.nx.all&query=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&oquery=%ED%8B%B0%EB%B9%99+%EC%B6%94%EC%B2%9C+%EC%98%81%ED%99%94&tqi=iP75klqo1e8ssO4UwQhssssssud-170148"
urls = [netflix_url, watcha_url, tving_url]
headers = ["제목", "국가", "장르", "평점", "개봉연도", "해시태그", "OTT"]


rawData_workbook = openpyxl.Workbook()
rawData_sheet = rawData_workbook.active
rawData_sheet.title = "RawData"
for col, header in enumerate(headers, 1):
    rawData_sheet.cell(1, col, header)
    rawData_sheet.cell(1, col).fill = PatternFill(start_color = "B7F0B1", fill_type= "solid")

chrome_options = Options()
chrome_options.add_experimental_option("detach", True) 
# chrome_options.add_argument("--headless")          ## 백그라운드 실행
driver = webdriver.Chrome(options=chrome_options)

for url in urls :
    driver.get(url)
    time.sleep(0.1)

    # 국가 선택
    for c in range(2,9) : 
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > a > span.menu._text").click()
        driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger2 > div > div > div > div > div > ul > li:nth-child(" + str(c) + ") > a").click()
        time.sleep(0.1)
        
    # 장르 선택
        for g in range(2, 37) : 
            driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > a > span.menu._text").click()
            driver.find_element(By.CSS_SELECTOR, "#main_pack > section.sc_new.cs_common_module.case_list.color_5._cs_contents_recommendation > div.cm_content_wrap > div > div > div.cm_tap_area > div > div > ul > li.tab._select_trigger3 > div > div > div > div > div > ul > li:nth-child(" + str(g) + ") > a").click()
            time.sleep(0.1)

            movieCountry = driver.find_element(By.XPATH, "//*[@id=\"main_pack\"]/section[1]/div[2]/div/div/div[1]/div/div/ul/li[2]/a").text
            movieGenre = driver.find_element(By.XPATH, "//*[@id=\"main_pack\"]/section[1]/div[2]/div/div/div[1]/div/div/ul/li[3]/a").text

            pageIdx = 1
            while True :
                for i in range(1, 9) : 
                    try :
                        movieName = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/strong/a").text
                    except :
                        continue
                
                    try :
                        a = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/div[2]/span[3]/span").text
                        if "." in a :
                            avgScore = a
                        else :
                            avgScore = "0.0"
                    except :
                        avgScore = "0.0"

                    try :
                        y = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/div[2]/span[2]").text
                        if (len(y) == 4 and "." not in y) :
                            year = y
                        elif "." in y :
                            avgScore = y
                            year = "0"
                        else :
                            year = "0"
                    except :
                        year = "0"

                    try :
                        h = driver.find_element(By.XPATH, "//*[@id=\"mflick\"]/div/div/ul["+str(pageIdx)+"]/li["+str(i)+"]/div[3]").text
                        if "#" in h :
                            hashTag = h
                        else :
                            hashTag = "태그정보X"
                    except :
                        hashTag = "태그정보X"

                    movieInfo = [movieName, movieCountry, movieGenre, avgScore, year, hashTag]

                    if url == netflix_url :
                        movieInfo.append("Netflix")

                    elif url == watcha_url :
                        movieInfo.append("Watcha")

                    else :
                        movieInfo.append("Tving")

                    print(movieInfo)
                    rawData_sheet.append(movieInfo)
                    time.sleep(0.1)

                try: 
                    next = driver.find_element( By.XPATH, '//*[@id="main_pack"]/section[1]/div[2]/div/div/div[3]/div/a[2]')
                    nowpage = driver.find_element(By.CLASS_NAME, 'npgs_now._current').text
                    nextpage = driver.find_element(By.CLASS_NAME, '_total').text
                    next.click()
                    pageIdx += 1
                    if int(nowpage) == int(nextpage):
                        break
                except :    
                    break
    
rawData_workbook.save("C:/Users/Public/Documents/RawData.xlsx")
driver.quit()
